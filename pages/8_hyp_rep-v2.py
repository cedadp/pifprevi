import streamlit as st
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Module 1 — Correspondances", layout="wide")
st.title("✈️ Module 1 — Calcul des taux de correspondance")

uploaded = st.file_uploader("Déposez le fichier Excel source", type=["xlsx"])

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
TERMINAL_MAP_APPORT = {
    "T1": "ABCDT1", "T2A": "ABCDT1", "T2B": "ABCDT1",
    "T2C": "ABCDT1", "T2D": "ABCDT1",
    "T2F": "C2F", "T2G": "C2G"
}
TERMINAL_MAP_EMPORT = TERMINAL_MAP_APPORT.copy()

SALLE_MAP_APPORT = {
    "C2E-JETE": "salle_K", "C2E-S3": "salle_L", "C2E_S4": "salle_M"
}
SALLE_MAP_EMPORT = {
    "C2E-JETE": "salle_K", "C2E-S3": "salle_L", "C2E_S4": "Salle_M"
}

ORIGINES  = ["salle_K", "salle_L", "salle_M", "C2F", "C2G"]
DESTINATIONS = ["ABCDT1", "C2F", "C2G", "salle_K", "salle_L", "Salle_M"]

ZONES_GEO = [
    "Métropole", "Schengen", "U.E. hors M & S",
    "Afrique du Nord", "Amérique du Nord", "Autre Afrique",
    "Autre Europe", "DOM TOM", "Extrême Orient",
    "Moyen Orient", "Amérique Centre + Sud"
]

# Faisceaux calculés réellement (uniquement pour origines C2F / C2G)
FAISCEAUX_CALCULES = ["Métropole", "Schengen", "U.E. hors M & S"]

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def map_terminal(val, mapping):
    return mapping.get(str(val).strip(), str(val).strip())

def map_salle(val, mapping):
    return mapping.get(str(val).strip(), str(val).strip())

def get_origine(row):
    """Détermine l'origine : salle si C2E, sinon terminal apport."""
    term = str(row.get("Vol apport - Terminal", "")).strip()
    salle = str(row.get("Vol apport - Ressources - Salle de débarquement", "")).strip()
    if term in ("C2F", "C2G") or map_terminal(term, TERMINAL_MAP_APPORT) in ("C2F", "C2G"):
        mapped_term = map_terminal(term, TERMINAL_MAP_APPORT)
        # Si salle renseignée et appartient aux salles C2E → origine = salle
        mapped_salle = SALLE_MAP_APPORT.get(salle, salle)
        if mapped_salle in ("salle_K", "salle_L", "salle_M"):
            return mapped_salle
        return mapped_term
    # Pour ABCDT1 : origine = terminal
    return map_terminal(term, TERMINAL_MAP_APPORT)

def get_destination(row):
    term = str(row.get("Vol emport - Terminal", "")).strip()
    salle = str(row.get("Vol emport - Ressources - Salle d'embarquement", "")).strip()
    mapped_term = map_terminal(term, TERMINAL_MAP_EMPORT)
    if mapped_term in ("C2F", "C2G"):
        mapped_salle = SALLE_MAP_EMPORT.get(salle, salle)
        if mapped_salle in ("salle_K", "salle_L", "Salle_M"):
            return mapped_salle
        return mapped_term
    return mapped_term

# ─────────────────────────────────────────────
# TRAITEMENT PRINCIPAL
# ─────────────────────────────────────────────
if uploaded:
    df_raw = pd.read_excel(uploaded)
    st.subheader("Aperçu des données brutes")
    st.dataframe(df_raw.head(20), use_container_width=True)

    df = df_raw.copy()

    # --- Renommage terminaux apport
    if "Vol apport - Terminal" in df.columns:
        df["Vol apport - Terminal"] = df["Vol apport - Terminal"].apply(
            lambda x: map_terminal(x, TERMINAL_MAP_APPORT))

    # --- Renommage salles apport
    sal_col_ap = "Vol apport - Ressources - Salle de débarquement"
    if sal_col_ap in df.columns:
        df[sal_col_ap] = df[sal_col_ap].apply(lambda x: SALLE_MAP_APPORT.get(str(x).strip(), str(x).strip()))

    # --- Renommage terminaux emport
    if "Vol emport - Terminal" in df.columns:
        df["Vol emport - Terminal"] = df["Vol emport - Terminal"].apply(
            lambda x: map_terminal(x, TERMINAL_MAP_EMPORT))

    # --- Renommage salles emport
    sal_col_em = "Vol emport - Ressources - Salle d'embarquement"
    if sal_col_em in df.columns:
        df[sal_col_em] = df[sal_col_em].apply(lambda x: SALLE_MAP_EMPORT.get(str(x).strip(), str(x).strip()))

    # --- Filtre même jour AIBT / AOBT
    if "AIBT" in df.columns and "AOBT" in df.columns:
        df["AIBT"] = pd.to_datetime(df["AIBT"], errors="coerce")
        df["AOBT"] = pd.to_datetime(df["AOBT"], errors="coerce")
        df = df[df["AIBT"].dt.date == df["AOBT"].dt.date]

    # --- Origine / Destination
    df["origine"]     = df.apply(get_origine, axis=1)
    df["destination"] = df.apply(get_destination, axis=1)

    # --- Faisceau géographique
    faisceau_col = None
    for c in df.columns:
        if "faisceau" in c.lower():
            faisceau_col = c
            break

    st.subheader("Données transformées")
    st.dataframe(df.head(30), use_container_width=True)

    # ─────────────────────────────────────────────
    # MATRICES GLOBALES (sans plage)
    # ─────────────────────────────────────────────
    df_od = df[df["origine"].isin(ORIGINES) & df["destination"].isin(DESTINATIONS)]

    count_matrix = pd.crosstab(df_od["origine"], df_od["destination"])
    rate_matrix  = count_matrix.div(count_matrix.sum(axis=1), axis=0).fillna(0)

    for orig in ORIGINES:
        if orig not in count_matrix.index:
            count_matrix.loc[orig] = 0
        if orig not in rate_matrix.index:
            rate_matrix.loc[orig] = 0
    for dest in DESTINATIONS:
        if dest not in count_matrix.columns:
            count_matrix[dest] = 0
        if dest not in rate_matrix.columns:
            rate_matrix[dest] = 0

    count_matrix = count_matrix.reindex(index=ORIGINES, columns=DESTINATIONS, fill_value=0)
    rate_matrix  = rate_matrix.reindex(index=ORIGINES, columns=DESTINATIONS, fill_value=0)

    st.subheader("Matrice des taux globaux")
    st.dataframe(rate_matrix.style.format("{:.4f}").background_gradient(cmap="YlGn"), use_container_width=True)

    # ─────────────────────────────────────────────
    # ANALYSE PAR PLAGE
    # ─────────────────────────────────────────────
    plage_col = None
    for c in df.columns:
        if "plage" in c.lower():
            plage_col = c
            break

    od_sheets = {}  # {sheet_name: DataFrame à écrire}

    if plage_col:
        all_plages = sorted(df[plage_col].dropna().unique())
        df_od_plage = df[df["origine"].isin(ORIGINES) & df["destination"].isin(DESTINATIONS)].copy()

        for orig in ORIGINES:
            for dest in DESTINATIONS:
                sheet_name = f"{orig}_{dest}"[:31]
                rows = []
                for plage in all_plages:
                    mask_plage = df_od_plage[plage_col] == plage
                    pax_orig_plage = df_od_plage[mask_plage & (df_od_plage["origine"] == orig)].shape[0]
                    pax_od = df_od_plage[
                        mask_plage &
                        (df_od_plage["origine"] == orig) &
                        (df_od_plage["destination"] == dest)
                    ].shape[0]

                    taux_global = round(pax_od / pax_orig_plage, 4) if pax_orig_plage > 0 else 0

                    row = {"plage": plage, "taux_global": taux_global}

                    # Calcul par faisceau — uniquement si origine C2F ou C2G
                    if orig in ("C2F", "C2G") and faisceau_col:
                        for fais in FAISCEAUX_CALCULES:
                            mask_fais = df_od_plage[faisceau_col].astype(str).str.strip() == fais
                            pax_orig_fais = df_od_plage[
                                mask_plage & mask_fais & (df_od_plage["origine"] == orig)
                            ].shape[0]
                            pax_od_fais = df_od_plage[
                                mask_plage & mask_fais &
                                (df_od_plage["origine"] == orig) &
                                (df_od_plage["destination"] == dest)
                            ].shape[0]
                            row[fais] = round(pax_od_fais / pax_orig_fais, 4) if pax_orig_fais > 0 else 0
                    else:
                        for fais in FAISCEAUX_CALCULES:
                            row[fais] = None  # vide pour les autres origines

                    rows.append(row)

                od_sheets[sheet_name] = {"orig": orig, "dest": dest, "data": rows}

        # Aperçu streamlit
        st.subheader("📊 Aperçu par OD et plage")
        od_keys = list(od_sheets.keys())
        selected_od = st.selectbox("Choisir une OD", od_keys)
        if selected_od:
            preview_data = od_sheets[selected_od]["data"]
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True)

    # ─────────────────────────────────────────────
    # EXPORT EXCEL
    # ─────────────────────────────────────────────
    output = io.BytesIO()
    wb = openpyxl.Workbook()

    # Onglet données transformées
    ws_data = wb.active
    ws_data.title = "Données transformées"
    cols = list(df.columns)
    ws_data.append(cols)
    for _, row in df.iterrows():
        ws_data.append([row[c] for c in cols])

    # Onglet matrice taux
    ws_rate = wb.create_sheet("Matrice taux globaux")
    ws_rate.append([""] + DESTINATIONS)
    for orig in ORIGINES:
        ws_rate.append([orig] + [round(rate_matrix.loc[orig, dest], 4)
                                  if orig in rate_matrix.index and dest in rate_matrix.columns else 0
                                  for dest in DESTINATIONS])

    # Onglets OD par plage
    if plage_col:
        for sheet_name, od_info in od_sheets.items():
            ws_od = wb.create_sheet(sheet_name)
            orig = od_info["orig"]
            data_rows = od_info["data"]

            # Ligne 1 : en-têtes
            ws_od.cell(row=1, column=1, value="heure")
            for j, zone in enumerate(ZONES_GEO):
                ws_od.cell(row=1, column=j + 2, value=zone)

            # Lignes de données
            for i, row in enumerate(data_rows):
                r = i + 2
                ws_od.cell(row=r, column=1, value=row["plage"])

                # Colonne B = Métropole
                metro_val = row["Métropole"] if orig in ("C2F", "C2G") and row.get("Métropole") is not None else row["taux_global"]
                ws_od.cell(row=r, column=2, value=metro_val)

                # Colonne C = Schengen
                schengen_val = row.get("Schengen") if orig in ("C2F", "C2G") else None
                ws_od.cell(row=r, column=3, value=schengen_val)

                # Colonne D = U.E. hors M & S
                ue_val = row.get("U.E. hors M & S") if orig in ("C2F", "C2G") else None
                ws_od.cell(row=r, column=4, value=ue_val)

                # Colonnes E à L = autres faisceaux → vides pour tous
                for j in range(4, 11):
                    ws_od.cell(row=r, column=j + 2, value=None)

    wb.save(output)
    output.seek(0)

    st.download_button(
        label="⬇️ Télécharger le fichier Excel",
        data=output,
        file_name="analyse_correspondances.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
