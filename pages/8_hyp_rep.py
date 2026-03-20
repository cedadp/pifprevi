import streamlit as st
import pandas as pd
import io
import matplotlib.pyplot as plt
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Transformation fichier correspondances", layout="wide")
st.title("📊 Transformation & Analyse des correspondances aéroportuaires")

uploaded_file = st.file_uploader("📁 Déposez votre fichier Excel ici", type=["xlsx"])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)

    st.subheader("1️⃣ Aperçu du fichier original")
    st.dataframe(df.head(20), use_container_width=True)
    st.write(f"**Dimensions originales :** {df.shape[0]} lignes × {df.shape[1]} colonnes")

    # ─────────────────────────────────────────────
    # RENOMMAGE DES COLONNES
    # ─────────────────────────────────────────────
    col_mapping = {}
    for col in df.columns:
        col_strip = col.strip()
        col_lower = col_strip.lower()
        if "terminal" in col_lower and "apport" in col_lower:
            col_mapping[col] = "terminal_apport"
        elif "débarquement" in col_lower or "debarquement" in col_lower:
            col_mapping[col] = "salle_debarquement"
        elif "terminal" in col_lower and "emport" in col_lower:
            col_mapping[col] = "terminal_emport"
        elif "embarquement" in col_lower:
            col_mapping[col] = "salle_embarquement"
        elif "aibt" in col_lower:
            col_mapping[col] = "AIBT"
        elif "aobt" in col_lower:
            col_mapping[col] = "AOBT"
        elif "passager" in col_lower or "nombre" in col_lower:
            col_mapping[col] = "nb_passagers"
        elif "plage" in col_lower:
            col_mapping[col] = "plage"

    df.rename(columns=col_mapping, inplace=True)

    # ─────────────────────────────────────────────
    # TRANSFORMATIONS
    # ─────────────────────────────────────────────
    def transform_terminal_apport(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        if val in ["T1", "T2A", "T2B", "T2C", "T2D"]:
            return "ABCDT1"
        elif val == "T2F":
            return "C2F"
        elif val == "T2G":
            return "C2G"
        return val

    def transform_terminal_emport(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        if val in ["T1", "T2A", "T2B", "T2C", "T2D"]:
            return "ABCDT1"
        elif val == "T2F":
            return "C2F"
        elif val == "T2G":
            return "C2G"
        return val

    def transform_salle_debarquement(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        if val == "C2E-JETE":
            return "salle_K"
        elif val == "C2E-S3":
            return "salle_L"
        elif val in ["C2E-S4", "C2E_S4"]:
            return "salle_M"
        return val

    def transform_salle_embarquement(val):
        if pd.isna(val):
            return val
        val = str(val).strip()
        if val == "C2E-JETE":
            return "salle_K"
        elif val == "C2E-S3":
            return "salle_L"
        elif val in ["C2E-S4", "C2E_S4"]:
            return "Salle_M"
        return val

    if "terminal_apport" in df.columns:
        df["terminal_apport"] = df["terminal_apport"].apply(transform_terminal_apport)
    if "terminal_emport" in df.columns:
        df["terminal_emport"] = df["terminal_emport"].apply(transform_terminal_emport)
    if "salle_debarquement" in df.columns:
        df["salle_debarquement"] = df["salle_debarquement"].apply(transform_salle_debarquement)
    if "salle_embarquement" in df.columns:
        df["salle_embarquement"] = df["salle_embarquement"].apply(transform_salle_embarquement)

    # ─────────────────────────────────────────────
    # FILTRE MÊME JOUR
    # ─────────────────────────────────────────────
    nb_avant = len(df)
    if "AIBT" in df.columns and "AOBT" in df.columns:
        df["AIBT"] = pd.to_datetime(df["AIBT"], errors="coerce")
        df["AOBT"] = pd.to_datetime(df["AOBT"], errors="coerce")
        df = df[df["AIBT"].dt.date == df["AOBT"].dt.date]
    nb_apres = len(df)

    st.subheader("2️⃣ Données transformées et filtrées")
    st.dataframe(df, use_container_width=True)

    col1, col2, col3 = st.columns(3)
    col1.metric("Lignes originales", nb_avant)
    col2.metric("Lignes conservées", nb_apres)
    col3.metric("Lignes supprimées", nb_avant - nb_apres)

    if "nb_passagers" in df.columns:
        st.metric("Total passagers en correspondance", int(df["nb_passagers"].sum()))

    output1 = io.BytesIO()
    with pd.ExcelWriter(output1, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Transformé")
    output1.seek(0)
    st.download_button(
        label="⬇️ Télécharger le fichier transformé",
        data=output1,
        file_name="fichier_transforme.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # ═════════════════════════════════════════════
    # PARTIE 2 — TAUX DE CORRESPONDANCE GLOBAL
    # ═════════════════════════════════════════════
    st.markdown("---")
    st.header("📈 Taux de correspondance")

    def get_origine(row):
        t = str(row.get("terminal_apport", "")).strip()
        s = str(row.get("salle_debarquement", "")).strip()
        if t == "T2E":
            return s if s not in ["", "nan"] else t
        return t

    def get_destination(row):
        t = str(row.get("terminal_emport", "")).strip()
        s = str(row.get("salle_embarquement", "")).strip()
        if t == "T2E":
            return s if s not in ["", "nan"] else t
        return t

    df["origine"]     = df.apply(get_origine, axis=1)
    df["destination"] = df.apply(get_destination, axis=1)

    ORIGINES     = ["salle_K", "salle_L", "salle_M", "C2F", "C2G"]
    DESTINATIONS = ["ABCDT1", "C2F", "C2G", "salle_K", "salle_L", "Salle_M"]
    pax_col      = "nb_passagers" if "nb_passagers" in df.columns else None

    rows_matrix = []
    rows_taux   = []
    for orig in ORIGINES:
        df_orig    = df[df["origine"] == orig]
        total_orig = int(df_orig[pax_col].sum()) if pax_col else len(df_orig)

        row_pax  = {"Origine \\ Destination": orig, "Total arrivée": total_orig}
        row_taux = {"Origine \\ Destination": orig, "Total arrivée": total_orig}
        for dest in DESTINATIONS:
            df_od  = df_orig[df_orig["destination"] == dest]
            pax_od = int(df_od[pax_col].sum()) if pax_col else len(df_od)
            taux   = round(pax_od / total_orig * 100, 1) if total_orig > 0 else 0.0
            row_pax[dest]  = pax_od
            row_taux[dest] = taux
        rows_matrix.append(row_pax)
        rows_taux.append(row_taux)

    matrix_pax  = pd.DataFrame(rows_matrix).set_index("Origine \\ Destination")
    matrix_taux = pd.DataFrame(rows_taux).set_index("Origine \\ Destination")

    st.subheader("🔢 Matrice des passagers (nombre absolu)")
    st.dataframe(matrix_pax.style.background_gradient(cmap="Blues", subset=DESTINATIONS), use_container_width=True)

    st.subheader("📊 Matrice des taux de correspondance (%)")
    st.markdown("*Lecture : pour 100 passagers arrivant à l'origine, X% repartent vers la destination.*")
    styled_taux = (
        matrix_taux[DESTINATIONS]
        .style
        .background_gradient(cmap="RdYlGn", axis=None, vmin=0, vmax=100)
        .format("{:.1f}%")
    )
    st.dataframe(styled_taux, use_container_width=True)

    st.subheader("📉 Visualisation graphique des taux (%)")
    fig, ax = plt.subplots(figsize=(12, 5))
    x      = np.arange(len(ORIGINES))
    width  = 0.13
    colors = ["#4e79a7", "#f28e2b", "#e15759", "#76b7b2", "#59a14f", "#edc948"]

    for i, dest in enumerate(DESTINATIONS):
        vals = [matrix_taux.loc[orig, dest] if orig in matrix_taux.index else 0 for orig in ORIGINES]
        ax.bar(x + i * width, vals, width, label=dest, color=colors[i % len(colors)])

    ax.set_xlabel("Origine", fontsize=11)
    ax.set_ylabel("Taux de correspondance (%)", fontsize=11)
    ax.set_title("Taux de correspondance par origine et destination", fontsize=13, fontweight="bold")
    ax.set_xticks(x + width * (len(DESTINATIONS) - 1) / 2)
    ax.set_xticklabels(ORIGINES, fontsize=10)
    ax.legend(title="Destination", bbox_to_anchor=(1.01, 1), loc="upper left")
    ax.set_ylim(0, 110)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.grid(axis="y", alpha=0.3)
    plt.tight_layout()
    st.pyplot(fig)

    # ═════════════════════════════════════════════
    # PARTIE 3 — ANALYSE PAR PLAGE
    # ═════════════════════════════════════════════
    st.markdown("---")
    st.header("🕐 Analyse par plage horaire")

    ZONE_HEADERS = [
        "Métropole",
        "Schengen",
        "U.E. hors M & S",
        "Afrique du Nord",
        "Amérique du Nord",
        "Autre Afrique",
        "Autre Europe",
        "DOM TOM",
        "Extrême Orient",
        "Moyen Orient",
        "Amérique Centre + Sud",
    ]

    if "plage" not in df.columns:
        st.warning("⚠️ Colonne 'plage' introuvable dans le fichier. Veuillez vérifier le fichier source.")
    else:
        plages_disponibles = sorted(df["plage"].dropna().unique())
        st.write(f"**Plages détectées :** {plages_disponibles}")

        # Toutes les combinaisons origine → destination présentes dans les données
        od_pairs = (
            df.groupby(["origine", "destination"])[pax_col]
            .sum()
            .reset_index()
        )
        od_pairs = od_pairs[
            (od_pairs["origine"].isin(ORIGINES)) &
            (od_pairs["destination"].isin(DESTINATIONS))
        ]

        unique_pairs = list(od_pairs[["origine", "destination"]].itertuples(index=False, name=None))

        if not unique_pairs:
            st.warning("Aucune paire origine-destination trouvée avec les filtres actuels.")
        else:
            st.write(f"**Paires OD générées :** {len(unique_pairs)} onglets")

            # ── Aperçu Streamlit ──────────────────────
            st.subheader("Aperçu : taux par plage pour chaque paire OD")

            for orig, dest in unique_pairs:
                df_od      = df[(df["origine"] == orig) & (df["destination"] == dest)]
                total_orig_par_plage = df[df["origine"] == orig].groupby("plage")[pax_col].sum()

                plage_stats = []
                for p in plages_disponibles:
                    pax_od_p    = int(df_od[df_od["plage"] == p][pax_col].sum()) if pax_col else len(df_od[df_od["plage"] == p])
                    total_p     = int(total_orig_par_plage.get(p, 0))
                    taux_p      = round(pax_od_p / total_p * 100, 1) if total_p > 0 else 0.0
                    plage_stats.append({"Plage": p, "Pax OD": pax_od_p, "Total origine": total_p, "Taux (%)": taux_p})

                df_preview = pd.DataFrame(plage_stats).set_index("Plage")

                with st.expander(f"📌 {orig} → {dest}"):
                    st.dataframe(df_preview.style.format({"Taux (%)": "{:.1f}%"}).background_gradient(cmap="YlOrRd", subset=["Taux (%)"]), use_container_width=True)

            # ── Génération du fichier Excel ───────────
            st.subheader("📥 Export Excel par plage")

            wb = Workbook()
            wb.remove(wb.active)  # supprime la feuille vide par défaut

            for orig, dest in unique_pairs:
                # Nom de l'onglet : max 31 caractères (limite Excel)
                sheet_name = f"{orig}_{dest}"
                sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]

                ws = wb.create_sheet(title=sheet_name)

                # ── Ligne 1 : en-têtes ─────────────────
                ws["A1"] = "heure"
                for col_idx, zone in enumerate(ZONE_HEADERS, start=2):
                    ws.cell(row=1, column=col_idx, value=zone)

                # ── Calcul des taux par plage ──────────
                df_od               = df[(df["origine"] == orig) & (df["destination"] == dest)]
                total_orig_par_plage = df[df["origine"] == orig].groupby("plage")[pax_col].sum()

                for row_idx, p in enumerate(plages_disponibles, start=2):
                    pax_od_p = int(df_od[df_od["plage"] == p][pax_col].sum()) if pax_col else len(df_od[df_od["plage"] == p])
                    total_p  = int(total_orig_par_plage.get(p, 0))
                    taux_p   = round(pax_od_p / total_p * 100, 1) if total_p > 0 else 0.0

                    # Colonne A : label plage
                    ws.cell(row=row_idx, column=1, value=str(p))

                    # Colonne B : taux calculé
                    ws.cell(row=row_idx, column=2, value=taux_p)

                    # Colonnes C→L : même valeur recopiée
                    for col_idx in range(3, 2 + len(ZONE_HEADERS)):
                        ws.cell(row=row_idx, column=col_idx, value=taux_p)

            # ── Sauvegarde dans le buffer ──────────────
            output3 = io.BytesIO()
            wb.save(output3)
            output3.seek(0)

            st.download_button(
                label="⬇️ Télécharger l'analyse par plage (Excel)",
                data=output3,
                file_name="analyse_par_plage.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # ── Export global (toutes parties) ────────────
        st.markdown("---")
        output2 = io.BytesIO()
        with pd.ExcelWriter(output2, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Données transformées")
            matrix_pax.to_excel(writer, sheet_name="Matrice passagers")
            matrix_taux.to_excel(writer, sheet_name="Matrice taux (%)")
        output2.seek(0)
        st.download_button(
            label="⬇️ Télécharger l'analyse complète (Excel)",
            data=output2,
            file_name="analyse_correspondances.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("👆 Veuillez déposer un fichier Excel pour commencer l'analyse.")
