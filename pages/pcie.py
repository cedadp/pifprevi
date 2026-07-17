import streamlit as st
import pandas as pd
import pdfplumber
import io
from openpyxl import load_workbook


st.set_page_config(page_title="Concaténateur Prévisions Cies", layout="wide")

OUTPUT_COLS = ["ArrDep", "CieOpe", "NumVol", "EscDep", "EscArr",
               "DateLocaleMvt", "NbPaxCNT", "NbPaxTOT"]

# ---------------------------------------------------------------
# UTILITAIRES
# ---------------------------------------------------------------
def normalize_columns(df):
    df.columns = (
        df.columns.astype(str)
        .str.replace(r"\s+", " ", regex=True)
        .str.strip()
    )
    return df

def parse_lh_date(serie):
    d = pd.to_datetime(serie, format="%d.%m.%y", errors="coerce")
    d = d.fillna(pd.to_datetime(serie, errors="coerce", dayfirst=True))
    return d.dt.strftime("%d/%m/%Y")

def split_flight(serie):
    s = serie.astype(str).str.replace(r"\s+", "", regex=True).str.upper()
    cie = s.str.extract(r"^([A-Z0-9]{2})")[0]
    num = s.str.extract(r"^[A-Z0-9]{2}(\d+)")[0]
    return cie, num

def finalize_output(out):
    out = out.copy()
    out["NumVol"] = pd.to_numeric(out["NumVol"], errors="coerce").fillna(0).astype(int)
    for c in ["ArrDep", "CieOpe", "EscDep", "EscArr"]:
        out[c] = out[c].astype(str).str.strip()
    out["NbPaxCNT"] = pd.to_numeric(out["NbPaxCNT"], errors="coerce").fillna(0).astype(int)
    out["NbPaxTOT"] = pd.to_numeric(out["NbPaxTOT"], errors="coerce").fillna(0).astype(int)
    out = out[out["NbPaxTOT"] != 0]
    out = out[out["CieOpe"].notna() & (out["CieOpe"] != "") & (out["CieOpe"].str.lower() != "nan")]
    return out[OUTPUT_COLS]

# ---------------------------------------------------------------
# AF (mapping générique)
# ---------------------------------------------------------------
def transform_af(file, conf, label="AF"):
    df = pd.read_excel(file, sheet_name=conf["sheet"])
    df = normalize_columns(df)
    mapping = conf["mapping"]

    missing = [c for c in mapping if c not in df.columns]
    if missing:
        st.error(f"[{label}] Colonnes manquantes : {missing}")
        st.write(f"[{label}] Colonnes trouvées :", list(df.columns))
        return None

    out = pd.DataFrame()
    for src_col, dst_col in mapping.items():
        out[dst_col] = df[src_col]

    if conf.get("filter_cie"):
        out = out[out["CieOpe"].astype(str).str.strip() == conf["filter_cie"]]

    dcol = conf.get("date_col")
    if dcol and dcol in out.columns:
        d = pd.to_datetime(out[dcol], errors="coerce", dayfirst=True)
        out[dcol] = d.dt.strftime("%d/%m/%Y")

    return out


# ---------------------------------------------------------------
# EZ  (easyJet : .xls, ArrDep déduit via CDG, EJU/EZY)
# ---------------------------------------------------------------
def transform_ez(file):
    """easyJet : fichier .xls unique contenant arrivées ET départs."""
    raw = pd.read_excel(file, sheet_name="Sheet", header=5, engine="xlrd")
    raw = normalize_columns(raw)

    raw = raw[raw["FLT"].notna()]
    raw["DEP"] = raw["DEP"].astype(str).str.strip().str.upper()
    raw["ARR"] = raw["ARR"].astype(str).str.strip().str.upper()
    raw = raw[(raw["DEP"] != "") & (raw["ARR"] != "")]
    raw = raw[raw["DEP"].isin(["CDG"]) | raw["ARR"].isin(["CDG"])]

    flt = raw["FLT"].astype(str).str.replace(r"\s+", "", regex=True).str.upper()
    has_eju = flt.str.match(r"^EJU\d+")
    cie = has_eju.map({True: "EJU", False: "EZY"})
    num = flt.str.extract(r"(\d+)$")[0]

    out = pd.DataFrame({
        "ArrDep": ["A" if a == "CDG" else "D" for a in raw["ARR"]],
        "CieOpe": cie.values,
        "NumVol": pd.to_numeric(num, errors="coerce").fillna(0).astype(int).values,
        "EscDep": raw["DEP"].values,
        "EscArr": raw["ARR"].values,
        "DateLocaleMvt": pd.to_datetime(raw["DATE"], format="%d/%m/%y",
                                        errors="coerce").dt.strftime("%d/%m/%Y").values,
        "NbPaxCNT": 0,
        "NbPaxTOT": pd.to_numeric(raw["EXP"], errors="coerce").fillna(0).astype(int).values,
    })
    return out


# ---------------------------------------------------------------
# NH  (PDF unique -> inbound NH215 + outbound NH216)
# ---------------------------------------------------------------
def transform_nh(file, direction):
    """
    Parse NH PDF (All Nippon Airways) — texte brut, PAS de table alignée.
    Le PDF a 2 sections : ///NH215/// (arrivées) et ///NH216/// (départs).
    Les dates et les lignes pax sont dans des blocs séparés, à réapparier dans l'ordre.
    Ligne pax type : "48 16 115 179 100%76%79% 83%" -> TOTAL = dernier entier avant le 1er %.
    """
    import re

    if direction == 'inbound':
        section = 'NH215'
        cie_ope, esc_dep, esc_arr, arr_dep = 'NH', 'HND', 'CDG', 'A'
    else:  # outbound
        section = 'NH216'
        cie_ope, esc_dep, esc_arr, arr_dep = 'NH', 'CDG', 'HND', 'D'

    month_map = {
        'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04',
        'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
        'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
    }

    # 1) Texte brut
    text = ""
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            text += t + "\n"

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]

    date_re = re.compile(
        r"\((?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\)-(\d{2})-([A-Za-z]{3})-(\d{2})"
    )
    # ligne pax : au moins un pourcentage présent -> on prend le dernier entier AVANT le 1er %
    pct_re = re.compile(r"\d+\s*%")
    int_re = re.compile(r"\d+")

    dates = []
    totals = []

    for ln in lines:
        # a) dates (une ligne peut contenir 1 date)
        m = date_re.search(ln)
        if m:
            day, mon, yy = m.group(1), m.group(2).upper(), m.group(3)
            mm = month_map.get(mon)
            if mm:
                dates.append(f"{day}/{mm}/20{yy}")
            # une ligne peut mélanger date + pax : on continue quand même pour capter le total

        # b) lignes pax : contiennent un %
        pm = pct_re.search(ln)
        if pm:
            pax_part = ln[:pm.start()]          # tout ce qui précède le 1er %
            ints = int_re.findall(pax_part)
            if ints:
                total = int(ints[-1])           # dernier entier avant le % = TOTAL
                totals.append(total)

    # 3) Restreindre à la section demandée : le PDF liste NH215 puis NH216 (ou l'inverse).
    #    On repère l'ordre des marqueurs pour découper dates/totals par section.
    order = [s for s in re.findall(r"NH21[56]", text)]
    # fallback simple : les deux sections ont le même nombre de mouvements (32/32).
    n = len(totals)
    half = n // 2 if n else 0

    # Détermine quelle moitié correspond à la section voulue selon l'ordre d'apparition
    first_section = None
    for s in order:
        first_section = s
        break

    if half and first_section:
        if section == first_section:
            sel_totals = totals[:half]
        else:
            sel_totals = totals[half:]
    else:
        sel_totals = totals

    # dates : mêmes dates pour les 2 sections -> on prend une moitié équivalente
    if len(dates) >= 2 * half and half:
        sel_dates = dates[:half]
    else:
        sel_dates = dates[:len(sel_totals)]

    # 4) Appariement
    rows = []
    for d, tot in zip(sel_dates, sel_totals):
        if tot == 0:
            continue
        rows.append({
            'ArrDep': arr_dep,
            'CieOpe': cie_ope,
            'NumVol': section,
            'EscDep': esc_dep,
            'EscArr': esc_arr,
            'DateLocaleMvt': d,
            'NbPaxCNT': 0,
            'NbPaxTOT': tot,
        })

    df = pd.DataFrame(rows)
    return df if not df.empty else None

# ---------------------------------------------------------------
# LH
# ---------------------------------------------------------------
def transform_lh_inbound(file):
    df = pd.read_excel(file, header=0)
    df = normalize_columns(df)
    st.write("Colonnes LH détectées :", list(df.columns))
    df["Arr Date"] = df["Arr Date"].ffill()
    df = df[df["Flt Nbr"].notna() & (df["Flt Nbr"].astype(str).str.strip() != "")]
    cie, num = split_flight(df["Flt Nbr"])

    booked = pd.to_numeric(df.get("Booked PAX", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
    estim  = pd.to_numeric(df.get("Estimated PAX", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
    paxtot = booked.where(booked >= estim, estim).astype(int)

    return pd.DataFrame({
        "ArrDep": "A",
        "CieOpe": cie.values,
        "NumVol": num.values,
        "EscDep": df["Origin"].astype(str).str.strip().values,
        "EscArr": "CDG",
        "DateLocaleMvt": parse_lh_date(df["Arr Date"]).values,
        "NbPaxCNT": 0,
        "NbPaxTOT": paxtot.values,
    })



def get_visible_sheet(file):
    """Retourne le nom du premier onglet visible du classeur."""
    file.seek(0)
    wb = load_workbook(file, read_only=True)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    wb.close()
    file.seek(0)
    if not visible:
        raise ValueError("Aucun onglet visible trouvé")
    return visible[0]


def transform_lh_outbound(file):
    sheet = get_visible_sheet(file)
    df = pd.read_excel(file, sheet_name=sheet, header=0)
    df = normalize_columns(df)
    df = df[df["Flt Nbr"].notna() & (df["Flt Nbr"].astype(str).str.strip() != "")]
    cie, num = split_flight(df["Flt Nbr"])

    booked = pd.to_numeric(df.get("Booked PAX", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
    estim  = pd.to_numeric(df.get("Estimated PAX", pd.Series(0, index=df.index)), errors="coerce").fillna(0)
    paxtot = booked.where(booked >= estim, estim).astype(int)

    return pd.DataFrame({
        "ArrDep": "D",
        "CieOpe": cie.values,
        "NumVol": num.values,
        "EscDep": "CDG",
        "EscArr": df["Dest"].astype(str).str.strip().values,
        "DateLocaleMvt": parse_lh_date(df["Dep Date"]).values,
        "NbPaxCNT": 0,
        "NbPaxTOT": paxtot.values,
    })

# ---------------------------------------------------------------
# AI / EI
# ---------------------------------------------------------------
def _target_format(file, sheet_name, cie_fixe=None, cnt_present=True):
    df = pd.read_excel(file, sheet_name=sheet_name, header=0)
    df = normalize_columns(df)

    df = df[df["ArrDep"].astype(str).str.strip().isin(["A", "D"])]

    out = pd.DataFrame({
        "ArrDep": df["ArrDep"].astype(str).str.strip().values,
        "CieOpe": (cie_fixe if cie_fixe else df["CieOpe"].astype(str).str.strip()),
        "NumVol": pd.to_numeric(df["NumVol"], errors="coerce").fillna(0).astype(int).values,
        "EscDep": df["EscDep"].astype(str).str.strip().values,
        "EscArr": df["EscArr"].astype(str).str.strip().values,
        "DateLocaleMvt": pd.to_datetime(df["DateLocaleMvt"], errors="coerce")
                           .dt.strftime("%d/%m/%Y").values,
        "NbPaxCNT": (pd.to_numeric(df["NbPaxCNT"], errors="coerce").fillna(0).astype(int).values
                     if cnt_present else 0),
        "NbPaxTOT": pd.to_numeric(df["NbPaxTOT"], errors="coerce").fillna(0).astype(int).values,
    })
    return out

def transform_ai(file):
    return _target_format(file, "Masque Prévisions CDG", cie_fixe=None, cnt_present=True)

def transform_ei(file):
    return _target_format(file, "Masque Prévisions CDG", cie_fixe="EI", cnt_present=False)


# ---------------------------------------------------------------
# MK
# ---------------------------------------------------------------

def transform_mk(file):
    sheet = get_header_sheet(file, required_col="ArrDep")   # 1er onglet visible contenant 'ArrDep'
    df = pd.read_excel(file, sheet_name=sheet, header=0)
    df = normalize_columns(df)

    # NbPaxTOT obligatoire, NbPaxCNT vide -> 0
    df["NbPaxTOT"] = pd.to_numeric(df["NbPaxTOT"], errors="coerce").fillna(0).astype(int)
    df["NbPaxCNT"] = pd.to_numeric(df["NbPaxCNT"], errors="coerce").fillna(0).astype(int)

    # Exclusion des lignes à 0 pax (même critère que AF/LH)
    df = df[df["NbPaxTOT"] > 0]

    return pd.DataFrame({
        "ArrDep":        df["ArrDep"].astype(str).str.strip().values,
        "CieOpe":        df["CieOpe"].astype(str).str.strip().values,
        "NumVol":        df["NumVol"].values,
        "EscDep":        df["EscDep"].astype(str).str.strip().values,
        "EscArr":        df["EscArr"].astype(str).str.strip().values,
        "DateLocaleMvt": pd.to_datetime(df["DateLocaleMvt"]).dt.strftime("%d/%m/%Y").values,
        "NbPaxCNT":      df["NbPaxCNT"].values,
        "NbPaxTOT":      df["NbPaxTOT"].values,
    })


def get_header_sheet(file, required_col):
    """1er onglet visible dont la 1re ligne contient la colonne demandée."""
    from openpyxl import load_workbook
    file.seek(0)
    wb = load_workbook(file, read_only=True)
    target = None
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
        if required_col in header:
            target = ws.title
            break
    wb.close()
    file.seek(0)
    if target is None:
        raise ValueError(f"Aucun onglet visible avec la colonne '{required_col}'")
    return target



# ---------------------------------------------------------------
# DECLARATION DES SOURCES
# ---------------------------------------------------------------
SOURCES = {
    "AF": {
        "input_type": "excel",
        "label": "AF — Programme brut (Excel)",
        "sheet": "Programme brut",
        "mapping": {
            "A/D": "ArrDep",
            "Cie Ope": "CieOpe",
            "Num Vol": "NumVol",
            "Esc Dep": "EscDep",
            "Esc Arr": "EscArr",
            "Local Date": "DateLocaleMvt",
            "Pax CNT TOT": "NbPaxCNT",
            "PAX TOT": "NbPaxTOT",
        },
        "date_col": "DateLocaleMvt",
        "filter_cie": "AF",
        "af": True,
    },
    "LH_IN":  {"input_type": "excel", "label": "LH — Arrivées (inbound)", "custom": transform_lh_inbound},
    "LH_OUT": {"input_type": "excel", "label": "LH — Départs (outbound)", "custom": transform_lh_outbound},
    "AI":     {"input_type": "excel", "label": "AI — Air India (Masque Prévisions CDG)", "custom": transform_ai},
    "EI":     {"input_type": "excel", "label": "EI — Aer Lingus (Masque Prévisions CDG)", "custom": transform_ei},
    "EZ":     {"input_type": "excel", "label": "EZ — easyJet (EJU/EZY)", "custom": transform_ez},
    "NH":     {"input_type": "pdf",   "label": "NH — All Nippon (PDF unique HND↔CDG)", "custom": None},
    "MK":     {"input_type": "excel", "label": "MK — Air Madagascar (Masque Prévisions CDG)", "custom": transform_mk},
}

# ---------------------------------------------------------------
# INTERFACE
# ---------------------------------------------------------------
st.title("✈️ Concaténateur de prévisions compagnies")
st.markdown("Déposez les fichiers puis cliquez sur **GO**.")

excel_sources = {k: v for k, v in SOURCES.items() if v["input_type"] == "excel"}
pdf_sources   = {k: v for k, v in SOURCES.items() if v["input_type"] == "pdf"}
uploaded = {}

st.header("📁 Fichiers Excel")
for name, conf in excel_sources.items():
    uploaded[name] = st.file_uploader(conf.get("label", name),
                                      type=["xlsx", "xls"], key=f"file_{name}")

st.header("📄 Fichiers PDF")
for name, conf in pdf_sources.items():
    uploaded[name] = st.file_uploader(conf.get("label", name),
                                      type=["pdf"], key=f"file_{name}")

st.divider()





# ---------------------------------------------------------------
# APERÇU TCD (avant GO) — construit dès qu'un fichier est déposé
# ---------------------------------------------------------------
def build_preview_frames():
    """Rejoue les transformations sur les fichiers déjà déposés, sans bloquer sur erreur."""
    frames = []
    # Excel
    for name, conf in excel_sources.items():
        up = uploaded.get(name)
        if up is None:
            continue
        try:
            if conf.get("af"):
                res = transform_af(up, conf, label=name)
            else:
                res = conf["custom"](up)
            if res is not None:
                res = finalize_output(res)
                if not res.empty:
                    frames.append(res)
        except Exception:
            pass
    # PDF (NH : inbound + outbound)
    for name, conf in pdf_sources.items():
        up = uploaded.get(name)
        if up is None:
            continue
        try:
            if name == "NH":
                pdf_bytes = up.read()
                up.seek(0)  # remet le curseur pour le GO ensuite
                for d in ("inbound", "outbound"):
                    res = transform_nh(io.BytesIO(pdf_bytes), direction=d)
                    if res is not None:
                        res = finalize_output(res)
                        if not res.empty:
                            frames.append(res)
        except Exception:
            pass
    return frames


def render_tcd(df):
    """Affiche les 2 TCD par CieOpe (comme la capture)."""
    d = df.copy()
    d["_date"] = pd.to_datetime(d["DateLocaleMvt"], format="%d/%m/%Y", errors="coerce")

    # --- TCD 1 : Mvts / NbPaxTOT / NbPaxCNT ---
    tcd1 = d.groupby("CieOpe").agg(
        **{
            "Nombre de Mvts": ("CieOpe", "size"),
            "Somme de NbPaxTOT": ("NbPaxTOT", "sum"),
            "Somme de NbPaxCNT": ("NbPaxCNT", "sum"),
        }
    ).reset_index().rename(columns={"CieOpe": "Cies"})
    tcd1 = tcd1.sort_values("Cies").reset_index(drop=True)
    tot1 = pd.DataFrame([{
        "Cies": "Total général",
        "Nombre de Mvts": tcd1["Nombre de Mvts"].sum(),
        "Somme de NbPaxTOT": tcd1["Somme de NbPaxTOT"].sum(),
        "Somme de NbPaxCNT": tcd1["Somme de NbPaxCNT"].sum(),
    }])
    tcd1 = pd.concat([tcd1, tot1], ignore_index=True)

    # --- TCD 2 : Mvts / Date début / Date fin ---
    tcd2 = d.groupby("CieOpe").agg(
        **{
            "Nombre de mvts": ("CieOpe", "size"),
            "Date début période": ("_date", "min"),
            "Date fin période": ("_date", "max"),
        }
    ).reset_index().rename(columns={"CieOpe": "Cies"})
    tcd2 = tcd2.sort_values("Cies").reset_index(drop=True)
    tot2 = pd.DataFrame([{
        "Cies": "Total général",
        "Nombre de mvts": tcd2["Nombre de mvts"].sum(),
        "Date début période": d["_date"].min(),
        "Date fin période": d["_date"].max(),
    }])
    tcd2 = pd.concat([tcd2, tot2], ignore_index=True)
    for c in ["Date début période", "Date fin période"]:
        tcd2[c] = pd.to_datetime(tcd2[c]).dt.strftime("%d/%m/%Y")

    # Formatage : séparateur de milliers (espace)
    for c in ["Nombre de Mvts", "Somme de NbPaxTOT", "Somme de NbPaxCNT"]:
        tcd1[c] = tcd1[c].astype(int).map(lambda x: f"{x:,}".replace(",", " "))
    tcd2["Nombre de mvts"] = tcd2["Nombre de mvts"].astype(int).map(lambda x: f"{x:,}".replace(",", " "))

    # Affichage resserré, côte à côte
    col1, col2 = st.columns([1, 1], gap="medium")
    with col1:
        st.markdown("**📊 Pax**")
        st.dataframe(
            tcd1, 
            use_container_width=False, 
            hide_index=True, 
            width=520,
            column_config={
                "Cies": st.column_config.TextColumn("Cies", width="small"),
                "Nombre de Mvts": st.column_config.TextColumn("Nb Mvts", width="medium"),
                "Somme de NbPaxTOT": st.column_config.TextColumn("NbPaxTOT", width="medium"),
                "Somme de NbPaxCNT": st.column_config.TextColumn("NbPaxCNT", width="small"),
            }
        )
    with col2:
        st.markdown("**📅 Périodes**")
        st.dataframe(
            tcd2, 
            use_container_width=False, 
            hide_index=True, 
            width=500,
            column_config={
                "Cies": st.column_config.TextColumn("Cies", width="small"),
                "Nombre de mvts": st.column_config.TextColumn("Nb mvts", width="small"),
                "Date début période": st.column_config.TextColumn("Début", width="medium"),
                "Date fin période": st.column_config.TextColumn("Fin", width="medium"),
            }
        )


# Affichage automatique de l'aperçu dès qu'au moins un fichier est présent
if any(v is not None for v in uploaded.values()):
    st.subheader("👁️ Aperçu des TCD (avant intégration)")
    preview = build_preview_frames()
    if preview:
        render_tcd(pd.concat(preview, ignore_index=True))
    else:
        st.info("Fichiers détectés mais aucune donnée exploitable pour l'instant.")

st.divider()



if st.button("🚀 GO", type="primary", use_container_width=True):
    frames = []

    # --- Sources Excel ---
    for name, conf in excel_sources.items():
        up = uploaded.get(name)
        if up is None:
            continue
        try:
            if conf.get("af"):
                res = transform_af(up, conf, label=name)
            else:
                res = conf["custom"](up)
            if res is None:
                continue
            res = finalize_output(res)
            if not res.empty:
                frames.append(res)
                st.success(f"[{name}] {len(res)} lignes intégrées.")
            else:
                st.warning(f"[{name}] 0 ligne après filtrage.")
        except Exception as e:
            st.error(f"[{name}] Erreur : {e}")

    # --- Sources PDF (NH : 1 fichier -> inbound + outbound) ---
    for name, conf in pdf_sources.items():
        up = uploaded.get(name)
        if up is None:
            continue
        try:
            if name == "NH":
                pdf_bytes = up.read()
                res_in  = transform_nh(io.BytesIO(pdf_bytes), direction="inbound")
                res_out = transform_nh(io.BytesIO(pdf_bytes), direction="outbound")
                total = 0
                for res in (res_in, res_out):
                    if res is None:
                        continue
                    res = finalize_output(res)
                    if not res.empty:
                        frames.append(res)
                        total += len(res)
                if total > 0:
                    st.success(f"[NH] {total} lignes intégrées.")
                else:
                    st.warning("[NH] 0 ligne après filtrage.")
        except Exception as e:
            st.error(f"[{name}] Erreur : {e}")

    if not frames:
        st.warning("Aucune donnée valide fournie.")
    else:
        final = pd.concat(frames, ignore_index=True)
        st.subheader(f"📦 Fichier final ({len(final)} lignes)")
        st.dataframe(final.head(100))

        header = ";".join(OUTPUT_COLS)
        body = final.to_csv(sep=";", index=False, header=False, lineterminator="\n")
        csv_out = header + "\n" + body

        st.download_button(
            "💾 Télécharger le CSV",
            data=csv_out.encode("utf-8"),
            file_name="Previs_cies.csv",
            mime="text/csv",
        )
